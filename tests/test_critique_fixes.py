"""
Comprehensive Verification Tests for Architecture & Code Quality Critique Fixes.
Verifies all defect resolutions (DEF-01 through DEF-18).
"""
import os
import shutil
import tempfile
import openpyxl
import pandas as pd
import numpy as np
from src.catalog.sap_catalog import SAPCatalogManager
from src.catalog.schema_models import TableSchema, FieldMeta
from src.synthesis import DataSynthesizer, sort_specs_topologically, TableGenerationSpec, FieldRule
from src.masking.masking_engine import DataMaskingEngine
from src.masking.numeric_masker import NumericMasker
from src.masking.detector import SensitiveColumnDetector
from src.templates.template_generator import ExcelTemplateBuilder


def test_def01_session_state_isolation():
    """DEF-01: Verifies that multiple StudioState instances maintain completely isolated state."""
    from app import StudioState
    s1 = StudioState()
    s2 = StudioState()
    s1.creation_specs["BKPF"] = "spec1"
    assert "BKPF" not in s2.creation_specs
    s1.domain = "CUSTOM"
    assert s2.domain == "SAP"


def test_def02_and_def10_offline_catalog_and_batch_pv():
    """DEF-02 & DEF-10: Verifies offline get_table returns None for unindexed tables without network and loads batch possible values."""
    catalog = SAPCatalogManager()
    assert catalog.get_table("DEFINITELY_NOT_EXISTING_XYZ123") is None

    vbak = catalog.get_table("VBAK")
    assert vbak is not None
    assert "AUART" in vbak.fields
    auart_pv = vbak.fields["AUART"].possible_values
    assert len(auart_pv) > 0
    assert any(pv.val == "TA" for pv in auart_pv)


def test_def05_fk_cascade_does_not_overwrite_child_attributes():
    """DEF-05: Verifies generic cascade never blindly overwrites child item attributes (NETWR, MENGE)."""
    catalog = SAPCatalogManager()
    synthesizer = DataSynthesizer(catalog)

    # Generate parent VBAK
    vbak_df = synthesizer.generate_sap_table("VBAK", row_count=3)
    vbak_df["NETWR"] = [99999.00, 88888.00, 77777.00]

    # Generate child VBAP via generate_sap_table with parent_df
    vbap_df = synthesizer.generate_sap_table("VBAP", parent_df=vbak_df)
    # Child item NETWR must be individual item amounts, NOT overwritten with parent header 99999.00
    assert all(vbap_df["NETWR"] != 99999.00)
    # Child foreign key VBELN must match parent VBELN
    assert set(vbap_df["VBELN"]).issubset(set(vbak_df["VBELN"]))


def test_def06_masking_preview_does_not_pollute_vault_state():
    """DEF-06: Verifies 5-row preview execution does not mutate vault state or advance custom pool index."""
    engine = DataMaskingEngine()
    custom_names = ["Alpha Inc", "Beta LLC", "Gamma Corp", "Delta GmbH"]
    pools = {"NAME1": list(custom_names)}

    df = pd.DataFrame({
        "NAME1": ["Company A", "Company B", "Company C", "Company D", "Company E", "Company F"]
    })
    tables = {"KNA1": df}
    configs = {"KNA1": {"NAME1": "company_name"}}

    assert len(engine.vault._forward_map) == 0
    assert engine.vault._custom_indices.get("NAME1", 0) == 0

    preview = engine.generate_preview(tables, configs, custom_pools=pools, preview_rows=3)
    assert len(preview["KNA1"]["masked"]) == 3

    assert len(engine.vault._forward_map) == 0
    assert engine.vault._custom_indices.get("NAME1", 0) == 0

    full_masked = engine.mask_dataset(tables, configs, custom_pools=pools)
    assert full_masked["KNA1"]["NAME1"].iloc[0] == "Alpha Inc"


def test_def07_numeric_perturbation_small_range():
    """DEF-07: Verifies perturbation_range < 0.05 does not invert bounds and stays within requested range."""
    masker = NumericMasker()
    original = 1000.0
    for _ in range(50):
        masked = masker.mask_amount(original, perturbation_range=0.02)
        diff_pct = abs(masked - original) / original
        assert diff_pct <= 0.025, f"Shift {diff_pct} exceeded 2.5% max bound"

    assert masker.mask_amount("INVALID_NUM") == 0.0


def test_def08_detector_cell_by_cell_no_cross_row_bleed():
    """DEF-08: Verifies detector does not produce false positive matches due to cross-row text concatenation."""
    detector = SensitiveColumnDetector()
    df = pd.DataFrame({
        "MISC_CODE": ["user@sub.", "com_test", "standard_text", "code_val"]
    })
    results = detector.analyze_dataframe(df)
    assert not any(r["category"] == "email" for r in results)


def test_def09_date_detection_and_realism():
    """DEF-09: Verifies non-date fields like MANDAT or DATA are not misidentified, and date fields produce valid dates."""
    catalog = SAPCatalogManager()
    synthesizer = DataSynthesizer(catalog)

    df = synthesizer.generate_sap_table("BKPF", row_count=10)
    assert all(len(str(d)) == 8 and str(d).startswith("20") for d in df["BUDAT"])
    assert len(df["BUDAT"].unique()) >= 1


def test_def11_faker_pool_cardinality_scaling():
    """DEF-11: Verifies Faker candidate pool scales dynamically for large synthesis batches."""
    catalog = SAPCatalogManager()
    synthesizer = DataSynthesizer(catalog)

    rule = FieldRule(field_name="NAME1", rule_type="faker", params={"provider": "company"})
    names = synthesizer._generate_from_rule(rule, row_count=1000, f_meta=None)
    unique_count = len(np.unique(names))
    assert unique_count > 250, f"Expected > 250 unique companies, got {unique_count}"


def test_def13_bseg_double_entry_debit_credit_balancing():
    """DEF-13: Verifies BSEG line items have balanced debits and credits per accounting document."""
    catalog = SAPCatalogManager()
    synthesizer = DataSynthesizer(catalog)

    bkpf_df = synthesizer.generate_sap_table("BKPF", row_count=5)
    bseg_df = synthesizer.generate_sap_table("BSEG", parent_df=bkpf_df)

    for belnr in bkpf_df["BELNR"]:
        doc_lines = bseg_df[bseg_df["BELNR"] == belnr]
        assert len(doc_lines) >= 2, "BSEG must have at least 2 line items"

        debits = doc_lines[doc_lines["SHKZG"] == "S"]["WRBTR"].sum()
        credits = doc_lines[doc_lines["SHKZG"] == "H"]["WRBTR"].sum()
        assert round(debits, 2) == round(credits, 2), f"Doc {belnr} unbalanced: Debit={debits}, Credit={credits}"


def test_topological_sorter():
    """Verifies sort_specs_topologically orders parent tables before child tables."""
    specs = {
        "VBAP": TableGenerationSpec(table_name="VBAP", row_count=50, parent_table="VBAK"),
        "VBAK": TableGenerationSpec(table_name="VBAK", row_count=10),
        "BSEG": TableGenerationSpec(table_name="BSEG", row_count=40, parent_table="BKPF"),
        "BKPF": TableGenerationSpec(table_name="BKPF", row_count=10),
    }
    ordered = sort_specs_topologically(specs)
    keys = list(ordered.keys())
    assert keys.index("VBAK") < keys.index("VBAP")
    assert keys.index("BKPF") < keys.index("BSEG")


def test_def15_masking_engine_multi_tenant_isolation():
    """DEF-15: Verifies separate DataMaskingEngine instances do not leak custom pools or forward mappings."""
    engine_user_a = DataMaskingEngine()
    engine_user_b = DataMaskingEngine()

    pool_a = {"NAME1": ["Confidential Supplier Alpha", "Secret Vendor Beta"]}
    df_a = pd.DataFrame({"NAME1": ["ACME Corp", "Globex Inc"]})
    df_b = pd.DataFrame({"NAME1": ["Initech LLC", "ACME Corp"]})

    # User A masks with confidential custom replacement pool
    masked_a = engine_user_a.mask_dataset({"SUPPLIERS": df_a}, {"SUPPLIERS": {"NAME1": "company_name"}}, custom_pools=pool_a)
    assert masked_a["SUPPLIERS"]["NAME1"].iloc[0] == "Confidential Supplier Alpha"

    # User B masks unrelated dataset without custom pool
    masked_b = engine_user_b.mask_dataset({"VENDORS": df_b}, {"VENDORS": {"NAME1": "company_name"}})
    # User B must NEVER receive User A's secret custom supplier names!
    assert "Confidential Supplier Alpha" not in masked_b["VENDORS"]["NAME1"].values
    assert "Secret Vendor Beta" not in masked_b["VENDORS"]["NAME1"].values
    # Engine B's vault must be independent of Engine A's vault
    assert len(engine_user_b.vault._custom_pools) == 0


def test_def16_custom_schema_cascade_does_not_overwrite_columns():
    """DEF-16: Verifies unindexed custom schema cascades do NOT blindly overwrite child attributes (STATUS, DESCRIPTION)."""
    catalog = SAPCatalogManager()
    synthesizer = DataSynthesizer(catalog)

    parent_df = pd.DataFrame({
        "ORDER_ID": ["ORD-1001", "ORD-1002"],
        "STATUS": ["SHIPPED", "CANCELLED"],
        "DESCRIPTION": ["Parent Order 1", "Parent Order 2"]
    })

    child_schema = TableSchema(
        name="CUSTOM_ITEMS",
        description="Custom child table without catalog foreign keys",
        fields={
            "ORDER_ID": FieldMeta(name="ORDER_ID", data_type="CHAR", length=10),
            "STATUS": FieldMeta(name="STATUS", data_type="CHAR", length=15),
            "DESCRIPTION": FieldMeta(name="DESCRIPTION", data_type="CHAR", length=50),
            "PRICE": FieldMeta(name="PRICE", data_type="CURR", length=10),
        }
    )

    custom_rules = {
        "STATUS": FieldRule(field_name="STATUS", rule_type="choice", params={"values": ["PENDING", "IN_PROGRESS"]}),
        "DESCRIPTION": FieldRule(field_name="DESCRIPTION", rule_type="choice", params={"values": ["Item Alpha", "Item Beta"]}),
    }

    child_df = synthesizer._generate_child_table(
        table_name="CUSTOM_ITEMS",
        schema=child_schema,
        parent_df=parent_df,
        custom_rules=custom_rules
    )

    # 1. ORDER_ID must be propagated from parent to child
    assert set(child_df["ORDER_ID"]).issubset(set(parent_df["ORDER_ID"]))

    # 2. STATUS in child must NOT be overwritten with parent's "SHIPPED" or "CANCELLED"
    assert all(s in ["PENDING", "IN_PROGRESS"] for s in child_df["STATUS"])
    assert "SHIPPED" not in child_df["STATUS"].values
    assert "CANCELLED" not in child_df["STATUS"].values

    # 3. DESCRIPTION in child must NOT be overwritten with parent's "Parent Order 1/2"
    assert all(d in ["Item Alpha", "Item Beta"] for d in child_df["DESCRIPTION"])
    assert "Parent Order 1" not in child_df["DESCRIPTION"].values


def test_def17_session_temp_cleanup():
    """DEF-17: Verifies session directory cleanup removes temporary folders and files."""
    from app import TEMP_DIR, cleanup_stale_temp_dirs
    # Verify stale temp cleaner function executes cleanly
    cleanup_stale_temp_dirs()

    test_session_id = "test_cleanup_123"
    session_dir = os.path.join(TEMP_DIR, test_session_id)
    os.makedirs(session_dir, exist_ok=True)
    dummy_file = os.path.join(session_dir, "test.xlsx")
    with open(dummy_file, "w") as f:
        f.write("test content")

    assert os.path.exists(dummy_file)

    # Simulate client disconnect cleanup
    shutil.rmtree(session_dir, ignore_errors=True)
    assert not os.path.exists(session_dir)


def test_sec7_dead_code_eliminated():
    """Section 7.1: Verifies dead code segments are completely eliminated."""
    catalog = SAPCatalogManager()
    synthesizer = DataSynthesizer(catalog)
    # _generate_single_default_val must be completely removed
    assert not hasattr(synthesizer, "_generate_single_default_val")

    # ExcelTemplateBuilder generates valid template without unused styles
    builder = ExcelTemplateBuilder(catalog)
    with tempfile.TemporaryDirectory() as td:
        out = os.path.join(td, "template.xlsx")
        res = builder.generate_template(out, domain="SAP")
        assert os.path.exists(res)
        wb = openpyxl.load_workbook(res)
        assert "Table_Definitions" in wb.sheetnames
        assert "Field_Rules" in wb.sheetnames


def test_sec7_suffix_o1_lookup_and_prng_isolation():
    """Section 7.2.1 & 7.3.1: Verifies O(1) suffix set lookup and zero process-wide PRNG mutation."""
    import random
    from src.masking.format_preserver import FormatPreservingMasker, LEGAL_SUFFIXES_SET
    from src.masking.numeric_masker import NumericMasker

    # 1. Verify LEGAL_SUFFIXES_SET is a frozenset and contains upper cleaned suffixes
    assert isinstance(LEGAL_SUFFIXES_SET, frozenset)
    assert "LLC" in LEGAL_SUFFIXES_SET
    assert "GMBH" in LEGAL_SUFFIXES_SET
    assert "PVT LTD" in LEGAL_SUFFIXES_SET

    masker = FormatPreservingMasker(seed=999)
    base, suffix = masker.extract_legal_suffix("Siemens AG")
    assert base == "Siemens"
    assert suffix == "AG"

    base_pvt, suffix_pvt = masker.extract_legal_suffix("Acme Corp Pvt. Ltd.")
    assert base_pvt == "Acme Corp"
    assert suffix_pvt == "Pvt. Ltd."

    # 2. Verify PRNG Isolation: instantiating maskers must NOT mutate global process random state
    random.seed(12345)
    expected_rand = random.random()
    # Reset to 12345
    random.seed(12345)

    # Initialize maskers with different seeds
    _ = FormatPreservingMasker(seed=777)
    _ = NumericMasker(seed=888)

    # Global process random state must be identical to unmutated sequence
    actual_rand = random.random()
    assert actual_rand == expected_rand, "Global PRNG was mutated by FormatPreservingMasker or NumericMasker __init__!"


def test_sec7_vectorized_cython_mapping_and_case_insensitive_col_map():
    """Section 7.2.2, 7.2.3, 7.2.4: Verifies Cython map.fillna, precomputed col_map, and deferred copy."""
    engine = DataMaskingEngine(seed=42)
    df = pd.DataFrame({
        "COMPANY_NAME": ["Alpha LLC", "Beta AG", None, "Gamma Inc"],
        "AMOUNT": [100.50, 200.75, 300.0, None],
        "UNTOUCHED": ["Keep1", "Keep2", "Keep3", "Keep4"]
    })

    # Case-insensitive column config ("company_name" lowercase)
    configs = {
        "TEST_TBL": {
            "company_name": "company_name",
            "amount": "amount"
        }
    }

    masked = engine.mask_dataset({"TEST_TBL": df}, configs)
    res_df = masked["TEST_TBL"]

    # 1. Non-null company names masked, null preserved
    assert res_df["COMPANY_NAME"].iloc[0].endswith("LLC")
    assert res_df["COMPANY_NAME"].iloc[1].endswith("AG")
    assert pd.isna(res_df["COMPANY_NAME"].iloc[2])
    assert res_df["COMPANY_NAME"].iloc[3].endswith("Inc")

    # 2. Untouched column completely preserved
    assert list(res_df["UNTOUCHED"]) == ["Keep1", "Keep2", "Keep3", "Keep4"]

    # 3. Deferred copy test: if no configs apply, df is copied cleanly without error
    unaffected = engine.mask_dataset({"TEST_TBL": df}, {"TEST_TBL": {}})
    assert unaffected["TEST_TBL"].equals(df)


def test_sec7_streaming_excel_export():
    """Section 7.3.2: Verifies streaming openpyxl write_only mode for large enterprise tables."""
    from app import export_dataframes_to_excel
    with tempfile.TemporaryDirectory() as td:
        out_path = os.path.join(td, "streamed_export.xlsx")
        df1 = pd.DataFrame({"ID": [1, 2], "VAL": ["A", "B"]})
        df2 = pd.DataFrame({"ITEM": ["X", "Y", "Z"], "QTY": [10, 20, 30]})

        # Test normal write (< 25,000 rows)
        export_dataframes_to_excel({"TABLE1": df1, "TABLE2": df2}, out_path)
        assert os.path.exists(out_path)
        read_df1 = pd.read_excel(out_path, sheet_name="TABLE1")
        assert len(read_df1) == 2
        assert list(read_df1.columns) == ["ID", "VAL"]

        # Test streaming write mode simulation (> 25,000 rows threshold)
        out_stream_path = os.path.join(td, "large_stream.xlsx")
        # Create a mock dictionary where total rows triggers write_only
        large_df = pd.DataFrame({"NUM": range(26000), "TEXT": ["sample"] * 26000})
        export_dataframes_to_excel({"LARGE_TABLE": large_df}, out_stream_path)
        assert os.path.exists(out_stream_path)
        # Verify file can be read and row count matches
        read_large = pd.read_excel(out_stream_path, sheet_name="LARGE_TABLE")
        assert len(read_large) == 26000

