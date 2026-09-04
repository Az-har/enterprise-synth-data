"""
Excel Template Builder.
Generates structured .xlsx configuration templates with pre-populated examples and offline SAP reference guides.
"""
import os
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..catalog.sap_catalog import SAPCatalogManager


class ExcelTemplateBuilder:
    """Creates standardized enterprise specification workbooks."""

    def __init__(self, catalog_manager: Optional[SAPCatalogManager] = None):
        self.catalog = catalog_manager or SAPCatalogManager()

    def generate_template(self, output_path: str, domain: str = "SAP") -> str:
        """
        Creates a formatted Excel template.
        Domain: 'SAP' (pre-fills FI/SD tables) or 'CUSTOM' (blank starter structure).
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")  # Slate-800
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        sub_font = Font(name="Segoe UI", size=10)
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        hint_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        # -------------------------------------------------------------
        # SHEET 1: Table_Definitions
        # -------------------------------------------------------------
        ws_tables = wb.create_sheet(title="Table_Definitions")
        headers_1 = ["Table_Name", "Row_Count", "Parent_Table", "Description / Business Process"]
        ws_tables.append(headers_1)

        if domain.upper() == "SAP":
            ws_tables.append(["BKPF", 500, "", "Finance: Accounting Document Header"])
            ws_tables.append(["BSEG", "", "BKPF", "Finance: Accounting Document Items (1-4 items per BKPF)"])
            ws_tables.append(["VBAK", 250, "", "Sales: Sales Order Header"])
            ws_tables.append(["VBAP", "", "VBAK", "Sales: Sales Order Items (1-4 items per VBAK)"])
        else:
            ws_tables.append(["Customers", 100, "", "Customer Master entity"])
            ws_tables.append(["Orders", "", "Customers", "Order headers (1-3 orders per Customer)"])

        self._style_sheet(ws_tables, header_fill, header_font, sub_font, thin_border)

        # -------------------------------------------------------------
        # SHEET 2: Field_Rules
        # -------------------------------------------------------------
        ws_fields = wb.create_sheet(title="Field_Rules")
        headers_2 = ["Table_Name", "Field_Name", "Rule_Type", "Parameters / Values", "Notes / Business Meaning"]
        ws_fields.append(headers_2)

        if domain.upper() == "SAP":
            sample_rules = [
                ["BKPF", "BUKRS", "choice", "1000: 0.6, 2000: 0.4", "Company Codes: 60% Germany (1000), 40% US (2000)"],
                ["BKPF", "BLART", "choice", "KR: 0.7, SA: 0.3", "Document Types: 70% Vendor Invoice (KR), 30% G/L (SA)"],
                ["BKPF", "WAERS", "choice", "EUR: 0.7, USD: 0.3", "Currencies: 70% Euro, 30% US Dollar"],
                ["BSEG", "WRBTR", "range", "min: 50.0, max: 25000.0, decimals: 2", "Line Item Invoice Amount"],
                ["BSEG", "MWSKZ", "choice", "V1: 0.8, V2: 0.2", "Tax Codes: 80% Standard 19% (V1), 20% Reduced 7% (V2)"],
                ["BSEG", "ZLSPR", "choice", " : 0.9, A: 0.1", "Payment Block: 90% Free, 10% Blocked (A)"],
                ["VBAK", "VBELN", "sequence", "prefix: 5, start: 100000000, pad: 10", "10-digit Sales Document Number Sequence"],
                ["VBAK", "AUART", "choice", "OR: 0.8, ZOR: 0.2", "Order Types: 80% Standard (OR), 20% Rush (ZOR)"],
                ["VBAK", "NETWR", "range", "min: 500.0, max: 85000.0, decimals: 2", "Total Order Value"],
                ["VBAP", "MATNR", "sequence", "start: 100010, pad: 18", "18-digit zero-padded SAP Material Number (ALPHA)"],
                ["VBAP", "NETWR", "range", "min: 25.0, max: 20000.0, decimals: 2", "Item Net Value"]
            ]
        else:
            sample_rules = [
                ["Customers", "Country", "choice", "US: 0.5, DE: 0.3, UK: 0.2", "Country distribution"],
                ["Customers", "Company_Name", "faker", "company", "Generate realistic enterprise company name"],
                ["Orders", "Total_Amount", "range", "min: 100.0, max: 5000.0, decimals: 2", "Order total"]
            ]

        for row in sample_rules:
            ws_fields.append(row)

        self._style_sheet(ws_fields, header_fill, header_font, sub_font, thin_border)

        # -------------------------------------------------------------
        # SHEET 3: Reference_Catalog (For SAP)
        # -------------------------------------------------------------
        if domain.upper() == "SAP":
            ws_ref = wb.create_sheet(title="SAP_Reference_Catalog")
            ws_ref.append(["Table", "Field", "Description", "Data_Type", "Length", "Allowed Domain Values"])
            
            # Populate from catalog
            tables = self.catalog.list_tables()
            for t in tables[:8]:  # Top core tables
                t_schema = self.catalog.get_table(t["name"])
                if t_schema:
                    for f_name, f_meta in t_schema.fields.items():
                        vals_str = ", ".join([f"{pv.val} ({pv.desc})" for pv in f_meta.possible_values[:4]])
                        ws_ref.append([
                            t_schema.name,
                            f_name,
                            f_meta.description,
                            f_meta.data_type,
                            f_meta.length,
                            vals_str
                        ])
            self._style_sheet(ws_ref, header_fill, header_font, sub_font, thin_border)

        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        return output_path

    def _style_sheet(self, ws, header_fill, header_font, sub_font, border):
        """Applies auto-width and cell styling."""
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.font = sub_font
                    cell.alignment = Alignment(vertical="center")

                val_str = str(cell.value or "")
                max_len = max(max_len, len(val_str))

            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        ws.row_dimensions[1].height = 28
